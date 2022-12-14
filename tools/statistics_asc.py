import argparse
import logging
import math
import os
import shutil
import time
from multiprocessing import Pool, cpu_count
import openpyxl
from tqdm import tqdm

import numpy as np
import matplotlib.pyplot as plt

logger = logging.getLogger(__name__)
fh = logging.StreamHandler()
fh_formatter = logging.Formatter('%(asctime)s - %(levelname)s： %(message)s', datefmt='%H:%M:%S')
fh.setFormatter(fh_formatter)
logger.addHandler(fh)

other_list = ['gsensor', "camera"]     # 参与统计的其他字段

row_count = 3
col_count = 4


class Statistics:

    def __init__(self, log_path, save_path=None):
        self.log_path = log_path
        self.save_path = save_path
        self.file_name = os.path.split(log_path)[-1].split(".")[0]
        # self.pool = Pool(4)                         # 初始化进程池

        self.id_map = {}                                # CAN设备数据统计映射表
        self.record_data = []                           # 日志数据

        self.wb = openpyxl.Workbook()                                           # Excel文件对象
        self.excel_ws = self.wb.create_sheet(title=self.file_name, index=0)     # Excel工作表对象
        self.excel_row_pos = 1

        self.img_contain = 12

        self.init()

    def init(self):
        if not (self.save_path and os.path.exists(self.save_path) and os.path.isdir(self.save_path)):
            self.save_path = os.path.join(os.path.dirname(self.log_path), "asc信号接收统计图表")
            if not os.path.exists(self.save_path):
                os.makedirs(self.save_path)

        # 初始化表头
        excel_header = ["设备id", '平均值', '最小值', '最大值', '标准差']
        for i, h in enumerate(excel_header):
            self.excel_ws.cell(column=i+1, row=self.excel_row_pos, value=h)
        self.excel_row_pos += 1

    def run(self):
        print("start:", self.log_path)
        with open(self.log_path) as f:
            lines = f.readlines()

            for index, line in enumerate(lines):
                if line == '' or index < 3:
                    continue

                # 格式化日志数据
                cols = line.strip().split("	")
                if not self.id_map.get(cols[2]):
                    self.id_map[cols[2]] = [{"ts": float(cols[0]), "id": cols[2]}]
                else:
                    self.id_map[cols[2]].append({"ts": float(cols[0]), "id": cols[2]})

        # print(len(self.id_map))
        self.statistics_id()
        self.export_excel()
        # render([self.record_data], file_name=os.path.join(self.save_path, self.file_name))


    def statistics_id(self):
        """
            统计数据接收情况
        """
        id_list = list(self.id_map.keys())
        img_num = 1
        for i in range(0, len(id_list), self.img_contain):
            split_id = id_list[i: i + self.img_contain]
            render_list = [self.id_map[dev_id] for dev_id in split_id]

            over_count = len(render_list) % col_count
            if over_count != 0:
                over_list = render_list[:-over_count]
                if not over_list:
                    continue
                file_name = os.path.join(self.save_path, f'{self.file_name}_{img_num}.png')
                self.render(over_list, title=f"{self.file_name}_{img_num}", file_name=file_name)
                # self.pool.apply(render_can, args=(render_list[:-over_count], f"{can}_{img_num}", self.save_path))
                render_list = render_list[-over_count:]
                img_num += 1
                if not render_list:
                    continue
            file_name = os.path.join(self.save_path, f'{self.file_name}_{img_num}.png')
            self.render(render_list, title=f"{self.file_name}_{img_num}", file_name=file_name)
            # self.pool.apply(render_can, args=(render_list, f"{can}_{img_num}", self.save_path))
            img_num += 1

    def render(self, data_list, title="", file_name="统计图表.png"):
        """渲染设备的接收情况图表"""
        if not data_list:
            return
        if os.path.exists(file_name):
            logger.debug(f"已存在渲染数据：{file_name}")
            return

        # 设置主图像素跟大小
        can_id_count = len(data_list)
        render_col_count = can_id_count if can_id_count < col_count else col_count
        render_row_count = int(math.ceil(can_id_count / col_count))
        plt.rcParams['figure.dpi'] = 200
        plt.rcParams['figure.figsize'] = 10 * render_col_count, render_row_count * 6

        count = 0
        for can_id_data in data_list:
            data_count = len(can_id_data)
            if data_count <= 1:
                continue

            count += 1
            # 渲染子图图表
            plt.subplot(render_row_count, render_col_count, count)
            interval_list = []
            timestamp_list = []
            prev_timestamp = 0
            for data in can_id_data:
                timestamp_list.append(data.get("ts"))
                if not prev_timestamp:
                    prev_timestamp = data.get("ts")
                    continue

                # 计算接收间隔
                time_interval = data.get("ts") - prev_timestamp
                interval_list.append(time_interval)
                prev_timestamp = data.get("ts")

            # 求均值、最大值、最小值、标准差
            avg_interval = np.mean(interval_list)
            min_interval = np.min(interval_list)
            max_interval = np.max(interval_list)
            std_interval = np.std(interval_list)
            # 写入到表格对象中
            col_data = [can_id_data[0].get('id'), avg_interval, min_interval, max_interval, std_interval]
            for i, d in enumerate(col_data):
                self.excel_ws.cell(column=i+1, row=self.excel_row_pos, value=d)
            self.excel_row_pos += 1

            # 渲染
            plt.title(f"{'dev id:'+can_id_data[0].get('id')} count:({data_count}) time:{int(can_id_data[-1].get('ts') - can_id_data[0].get('ts'))}s std/avg:{'%.3f' % (std_interval/avg_interval)}")
            plt.xlabel("timestamp")
            plt.ylabel("interval(s)")
            avg_line = plt.axhline(y=avg_interval, color="r", linestyle="--", linewidth=1)
            min_line = plt.axhline(y=min_interval, color="g", linestyle="--", linewidth=1)
            max_line = plt.axhline(y=max_interval, color="y", linestyle="--", linewidth=1)
            plt.legend([avg_line, min_line, max_line], [f'avg:{"%.8f" % avg_interval}', f'min:{"%.8f" % min_interval}',
                                                        f'max:{"%.8f" % max_interval}'], bbox_to_anchor=(0, -0.23, 1, 2),
                       loc="lower left", mode="expand", borderaxespad=0, ncol=3)
            plt.scatter(timestamp_list[1:], interval_list, s=1, marker='.')  # s表示面积，marker表示图形

        plt.subplots_adjust(left=0, bottom=0, right=1, top=1, hspace=0.1, wspace=0.1)
        plt.tight_layout()
        plt.suptitle(title, fontsize="xx-large")
        plt.savefig(file_name, dpi=200, format='png')
        plt.close()

    def export_excel(self):
        excel_filename = os.path.join(self.save_path, self.file_name + ".xlsx")
        if os.path.exists(excel_filename):
            logger.debug(f"{excel_filename}已存在，跳过保存")
            self.wb.close()
        else:
            self.wb.save(excel_filename)


def asc_list_from_path(path):
    """
    从路径中自动遍历识别.asc文件路径列表
    @param path:
    @return:list
    """
    if not os.path.exists(path):
        logger.error(f"{path}路径不存在")
        return

    if os.path.isfile(path) and path.endswith(".asc"):
        return [path]
    elif os.path.isdir(path):
        log_list = []
        for f in os.listdir(path):
            children_path = os.path.join(path, f)
            if children_path.endswith(".asc"):
                log_list.append(children_path)
            elif os.path.isdir(children_path):
                log_list += asc_list_from_path(children_path)
        return log_list


if __name__ == "__main__":
    # 解析命令行参数
    parser = argparse.ArgumentParser()
    parser.add_argument('path', nargs='+', help='包含.asc文件的路径')
    parser.add_argument('--debug', action='store_true', help='调试模式', default=False)
    parser.add_argument('--save', '-s', help='保存统计图表文件夹路径（默认当前目录）')
    args = parser.parse_args()

    # 初始化日志输出等级
    if args.debug:
        logger.setLevel(logging.DEBUG)
    else:
        logger.setLevel(logging.WARNING)

    # 初始化保存路径
    if args.save and not os.path.exists(args.save):
        args.save = None

    log_path_list = []
    for path in args.path:
        if not os.path.exists(path):
            continue
        log_path_list += asc_list_from_path(path)

    logger.debug(f"wait statistics count: {len(log_path_list)}")
    for path in tqdm(log_path_list):
        task = Statistics(path, save_path=args.save)
        task.run()

    logger.warning("运行结束")

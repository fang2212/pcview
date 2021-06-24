import argparse
import logging
import math
import os
import time

import openpyxl
from tqdm import tqdm

import numpy as np
import matplotlib as mpl
import matplotlib.pyplot as plt

mpl.rcParams['agg.path.chunksize'] = 10000

logger = logging.getLogger(__name__)
fh = logging.StreamHandler()
fh_formatter = logging.Formatter('%(asctime)s - %(levelname)s： %(message)s', datefmt='%H:%M:%S')
fh.setFormatter(fh_formatter)
logger.addHandler(fh)

other_list = ['gsensor', "camera"]     # 参与统计的其他字段

row_count = 3
col_count = 4


class Statistics:

    def __init__(self, log_path, save_path=None, assign_ids=None):

        self.other_list = ['gsensor', "camera"]     # 不参与统计的字段
        self.log_path = log_path
        self.save_path = save_path
        self.log_camera_start_ts = 0
        self.log_camera_end_ts = 0
        self.log_start_ts = 0
        self.log_end_ts = 0
        self.log_long_ts = 0
        # self.pool = Pool(4)                         # 初始化进程池

        self.assign_ids = assign_ids                # 指定渲染id
        self.can_map = {}                           # CAN设备数据统计映射表
        self.device_map = {}                        # 设备数据统计映射表

        self.wb = openpyxl.Workbook()  # Excel文件对象
        self.excel_ws = self.wb.create_sheet(title="data", index=0)  # Excel工作表对象
        self.excel_row_pos = 1

        self.img_contain = 12

        self.init()

    def init(self):
        if not (self.save_path and os.path.exists(self.save_path) and os.path.isdir(self.save_path)):
            self.save_path = os.path.join(os.path.dirname(self.log_path), "log信号接收统计图表")
            if not os.path.exists(self.save_path):
                os.makedirs(self.save_path)

        # 初始化表头
        excel_header = ["设备名", "设备id", '平均值', '最小值', '最大值', '标准差']
        for i, h in enumerate(excel_header):
            self.excel_ws.cell(column=i + 1, row=self.excel_row_pos, value=h)
        self.excel_row_pos += 1

    def run(self):
        print("start:", self.log_path)
        with open(self.log_path) as f:
            lines = f.readlines()

            for line in lines:
                if line == '':
                    continue

                # 格式化日志数据
                cols = line.strip().split(" ")

                # 记录log开始结束的时间点
                ts = float(cols[0]) + float(cols[1]) / 1000000
                if self.log_start_ts == 0 or self.log_start_ts > ts:
                    self.log_start_ts = ts
                if self.log_end_ts < ts:
                    self.log_end_ts = ts

                if "CAN" in cols[2]:
                    self.can_collect(cols, ts)
                elif not ("voice_note" in cols[2] or "pinpoint" in cols[2]):
                    self.other_collect(cols, ts)

        self.log_long_ts = self.log_end_ts - self.log_start_ts
        logger.debug("视频开始时间：{} {}".format(self.log_camera_start_ts, time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(self.log_camera_start_ts))))
        logger.debug("视频结束时间：{} {}".format(self.log_camera_end_ts, time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(self.log_camera_end_ts))))
        logger.debug("log.txt开始时间：{} {}".format(self.log_start_ts, time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(
            self.log_start_ts))))
        logger.debug("log.txt结束时间：{} {}".format(self.log_end_ts,
                                           time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(self.log_end_ts))))
        logger.debug("log总时长跨度：{}".format(self.log_long_ts))
        self.statistics_can()
        self.statistics_other()

        self.export_excel()

    def can_collect(self, can_data, ts):
        data = {
            "ts": ts,
            "name": can_data[2],
            "id": can_data[3]
        }

        # 如果指定id 并当前id不在指定范围内
        if self.assign_ids and can_data[3] not in self.assign_ids:
            return

        # 更新CAN数据表
        if not self.can_map.get(can_data[2]):
            logger.debug(f"设备：{can_data[2]}")
            self.can_map[can_data[2]] = {}

        # 更新CAN id表
        if self.can_map[can_data[2]].get(can_data[3]):
            self.can_map[can_data[2]][can_data[3]].append(data)
        else:
            self.can_map[can_data[2]][can_data[3]] = [data]

    def other_collect(self, other_data, ts):
        if self.assign_ids:
            return

        data = {
            "ts": ts,
            "name": other_data[2],
        }
        self.log_end_ts = ts

        # 记录视频开始结尾时间
        if "camera" in other_data[2]:
            if self.log_camera_start_ts == 0 or self.log_camera_start_ts > ts:
                self.log_camera_start_ts = ts
            if self.log_camera_end_ts < ts:
                self.log_camera_end_ts = ts

        # 更新设备数据表
        if not self.device_map.get(other_data[2]):
            logger.debug(f"设备：{other_data[2]}")
            self.device_map[other_data[2]] = []

        self.device_map[other_data[2]].append(data)

    def statistics_other(self):
        render_list = []
        for device in self.device_map:
            render_list.append(self.device_map[device])
            file_name = os.path.join(self.save_path,
                                     f'{device.replace(".", "_")}_{time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime(self.device_map[device][0].get("ts")))}.png')
            self.render_can([self.device_map[device]], file_name=file_name, title=f"")
            self.device_map[device] = None

    def statistics_can(self):
        """
            统计CAN数据接收情况
            """
        for can in self.can_map:
            id_list = list(self.can_map[can].keys())
            img_num = 1
            for i in range(0, len(id_list), self.img_contain):
                split_can_id = id_list[i: i + self.img_contain]
                render_list = [self.can_map[can][can_id] for can_id in split_can_id]

                over_count = len(render_list) % col_count
                if over_count != 0 and len(render_list) > col_count:
                    over_list = render_list[:-over_count]
                    if not over_list:
                        continue
                    file_name = os.path.join(self.save_path, f'{can}_{img_num}_{time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime(over_list[0][0].get("ts")))}.png')
                    self.render_can(over_list, title=f"{can}_{img_num}", file_name=file_name)
                    # self.pool.apply(render_can, args=(render_list[:-over_count], f"{can}_{img_num}", self.save_path))
                    render_list = render_list[-over_count:]
                    img_num += 1
                    if not render_list:
                        continue
                file_name = os.path.join(self.save_path, f'{can}_{img_num}_{time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime(render_list[0][0].get("ts")))}.png')
                self.render_can(render_list, title=f"{can}", file_name=file_name)
                # self.pool.apply(render_can, args=(render_list, f"{can}_{img_num}", self.save_path))
                img_num += 1
            self.can_map[can] = None
            # pool.close()
            # pool.join()

    def render_can(self, data_list, title="", file_name="统计图表.png"):
        """渲染设备的接收情况图表"""
        if not data_list:
            return
        if os.path.exists(file_name):
            logger.debug(f"已存在渲染数据：{file_name}")
            return

        # 设置主图像素跟大小
        can_id_count = len(data_list)
        render_col_count = can_id_count if can_id_count < col_count else col_count
        render_row_count = int(math.ceil(can_id_count / col_count))
        plt.rcParams['figure.dpi'] = 200
        plt.rcParams['figure.figsize'] = 10 * render_col_count, render_row_count * 6

        count = 0
        has_draw = False
        for can_id_data in data_list:
            data_count = len(can_id_data)
            if data_count <= 1:
                continue

            # logger.debug("draw point count: {}, file:{}".format(data_count, file_name))
            count += 1
            # 渲染子图图表
            plt.subplot(render_row_count, render_col_count, count)
            interval_list = []
            timestamp_list = []
            prev_timestamp = 0
            for data in can_id_data:
                timestamp_list.append(data.get("ts"))
                if not prev_timestamp:
                    prev_timestamp = data.get("ts")
                    continue

                # 计算接收间隔
                time_interval = data.get("ts") - prev_timestamp
                interval_list.append(time_interval)
                prev_timestamp = data.get("ts")

            # 求均值、最大值、最小值、标准差
            avg_interval = np.mean(interval_list)
            min_interval = np.min(interval_list)
            max_interval = np.max(interval_list)
            std_interval = np.std(interval_list)

            # 将时间轴进行计算，通过减去最小值从0开始计算
            timestamp_list = [i-self.log_camera_start_ts for i in timestamp_list]

            # 写入到表格对象中
            col_data = [can_id_data[0].get('name', ""), can_id_data[0].get('id', ""), avg_interval, min_interval, max_interval, std_interval]
            for i, d in enumerate(col_data):
                self.excel_ws.cell(column=i + 1, row=self.excel_row_pos, value=d)
            self.excel_row_pos += 1

            # 渲染
            plt.title(f"{title +' id:'+can_id_data[0].get('id') if can_id_data[0].get('id') else ''} count:({data_count}) time:{int(self.log_long_ts)}s std/avg:{'%.3f' % (std_interval/avg_interval)}")
            plt.xlabel("timestamp")
            plt.ylabel("interval(s)")
            plt.xlim([self.log_start_ts - self.log_camera_start_ts, self.log_end_ts - self.log_camera_start_ts])
            avg_line = plt.axhline(y=avg_interval, color="r", linestyle="--", linewidth=1)
            min_line = plt.axhline(y=min_interval, color="g", linestyle="--", linewidth=1)
            max_line = plt.axhline(y=max_interval, color="y", linestyle="--", linewidth=1)
            plt.legend([avg_line, min_line, max_line], [f'avg:{"%.8f" % avg_interval}', f'min:{"%.8f" % min_interval}',
                                                        f'max:{"%.8f" % max_interval}'], bbox_to_anchor=(0, -0.23, 1, 2),
                       loc="lower left", mode="expand", borderaxespad=0, ncol=3)
            plt.plot(timestamp_list[1:], interval_list, marker='.', linewidth=1)  # s表示面积，marker表示图形
            has_draw = True

        if not has_draw:
            return

        plt.subplots_adjust(left=0, bottom=0, right=1, top=1, hspace=0.1, wspace=0.1)
        plt.tight_layout()
        # plt.suptitle(title, fontsize="xx-large")
        # print(file_name)
        # plt.show()
        plt.savefig(file_name, dpi=200, format='png')
        plt.clf()
        plt.close()

    def export_excel(self):
        excel_filename = os.path.join(self.save_path, "接收数据统计表.xlsx")
        if os.path.exists(excel_filename):
            logger.debug(f"{excel_filename}已存在，跳过保存")
            self.wb.close()
        else:
            self.wb.save(excel_filename)


def log_list_from_path(path):
    """
    从路径中自动遍历识别log.txt文件路径列表
    @param path:
    @return:list
    """
    if not os.path.exists(path):
        logger.error(f"{path}路径不存在")
        return

    if os.path.isfile(path) and os.path.split(path)[-1] == "log.txt":
        return [path]
    elif os.path.isdir(path):
        log_path = os.path.join(path, "log.txt")
        if os.path.exists(log_path):
            return [log_path]
        log_list = []
        for f in os.listdir(path):
            children_path = os.path.join(path, f)
            if os.path.isdir(children_path):
                log_list += log_list_from_path(children_path)
        return log_list


if __name__ == "__main__":
    # 解析命令行参数
    parser = argparse.ArgumentParser()
    parser.add_argument('path', nargs='+', help='包含log.txt的路径')
    parser.add_argument('--id', "-i", nargs='*', help='包含log.txt的路径')
    parser.add_argument('--debug', "-d", action='store_true', help='调试模式', default=False)
    parser.add_argument('--save', '-s', help='保存统计图表文件夹路径（默认当前目录）')
    args = parser.parse_args()

    # 初始化日志输出等级
    if args.debug:
        logger.setLevel(logging.DEBUG)
    else:
        logger.setLevel(logging.WARNING)

    # 初始化保存路径
    if args.save and not os.path.exists(args.save):
        args.save = None

    log_path_list = []
    for path in args.path:
        if not os.path.exists(path):
            continue
        log_path_list += log_list_from_path(path)

    logger.debug(f"wait statistics count: {len(log_path_list)}")
    for path in tqdm(log_path_list):
        task = Statistics(path, save_path=args.save, assign_ids=args.id)
        task.run()
        logger.warning("运行结束")
